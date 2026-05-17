import streamlit as st
import openpyxl
from google import genai
from google.genai import types
from pdf2image import convert_from_bytes
import io
import json
import os
import re

st.set_page_config(page_title="Multi-Vendor Comparative Analysis Engine", layout="wide")
st.title("📊 Multi-Vendor Quotation Matcher")
st.subheader("Upload your template and multiple vendor PDFs to automatically map names and material prices side-by-side.")

# Sidebar for Setup
with st.sidebar:
    st.header("Authentication")
    api_key = st.text_input("Enter Gemini API Key:", type="password")
    st.markdown("[Get a free API Key here](https://aistudio.google.com/)")

# File Uploaders
template_file = st.file_uploader("1. Upload Excel Template (.xlsx)", type=["xlsx"])
pfi_files = st.file_uploader("2. Upload ALL Vendor Quotations (Select multiple PDFs)", type=["pdf"], accept_multiple_files=True)

def convert_pdf_to_images(pdf_file):
    """Converts multi-page PDFs into clean images for layout scanning."""
    try:
        pdf_bytes = pdf_file.read()
        pdf_file.seek(0)
        return convert_from_bytes(pdf_bytes)
    except Exception as e:
        st.error(f"Failed to process pages for {pdf_file.name}: {e}")
        return []

def extract_all_vendors_simultaneously(all_document_packages, row_item_dict, api_key_str):
    """Sends all vendor invoices together to Gemini to align prices into parallel spreadsheet coordinates."""
    try:
        client = genai.Client(api_key=api_key_str)
        
        prompt_text = f"""You are an elite procurement analytics engine reviewing multiple vendor quotations simultaneously.
        
        Our Master Spreadsheet target lines (Format is "ROW_NUMBER": "EXACT MATERIAL DESCRIPTION"):
        {json.dumps(row_item_dict, indent=2)}
        
        YOUR INSTRUCTIONS:
        1. Look through each attached vendor quotation package individually.
        2. Extract the official Vendor Name from the letterhead/logo area of that specific package. Strip out any stray quotes.
        3. Match the items on that invoice to our Master Spreadsheet target list using flexible keyword matching (e.g., abbreviations like BLT, SS, MS, DIA match formal equivalents).
        4. Extract ONLY the numeric UNIT PRICE for that item. Ignore quantities, totals, or item serial numbers.
        5. If a vendor did not quote an item on our master list, completely omit its row number from that vendor's dictionary block.
        
        OUTPUT FORMATTING RULE:
        Return your findings strictly as a clean JSON object matching this schema layout. Group your answers by the exact 'file_index' provided to you. Do not include markdown code block syntax formatting or backslashes.
        
        {{
            "quotations": [
                {{
                    "file_index": 0,
                    "vendor_name": "Official Vendor Name 1",
                    "row_prices": {{
                        "9": 1500.00,
                        "12": 435.50
                    }}
                }}
            ]
        }}"""
        
        contents = [prompt_text]
        
        # Inject all separate document visual packages into the single data array stream
        for idx, package in enumerate(all_document_packages):
            contents.append(f"\n--- START OF VENDOR QUOTATION FILE INDEX {idx} (Filename: {package['name']}) ---")
            for img in package['images']:
                img_byte_arr = io.BytesIO()
                img.save(img_byte_arr, format='JPEG')
                contents.append(types.Part.from_bytes(data=img_byte_arr.getvalue(), mime_type="image/jpeg"))
            contents.append(f"--- END OF VENDOR QUOTATION FILE INDEX {idx} ---")
            
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=contents,
            config=types.GenerateContentConfig(response_mime_type="application/json")
        )
        
        return json.loads(response.text.strip())
    except Exception as e:
        st.error(f"Failed to communicate with master alignment engine: {e}")
        return {"quotations": []}

if st.button("🚀 Match & Populate Comparative Sheet") and template_file and pfi_files:
    if not api_key:
        st.warning("Please enter your Gemini API Key in the sidebar.")
    else:
        with st.spinner("Analyzing all vendor documents and running comparative grid alignment..."):
            wb = openpyxl.load_workbook(template_file)
            ws = wb.active
            
            # Read all master materials from Column C (Row 9 down)
            row_item_dict = {}
            for row in range(9, ws.max_row + 1):
                item_desc = ws.cell(row=row, column=3).value
                if item_desc:
                    clean_desc = re.sub(r'\s+', ' ', str(item_desc)).strip()
                    if "ITEM DESCRIPTION" in clean_desc.upper() or not clean_desc:
                        continue
                    row_item_dict[str(row)] = clean_desc
            
            st.info(f"Loaded {len(row_item_dict)} target material specifications from template.")
            
            # Convert up to 4 PDFs into structured visual packages
            all_packages = []
            for pfi in pfi_files[:4]:
                st.write(f"📷 Scanning text coordinates for document: **{pfi.name}**...")
                imgs = convert_pdf_to_images(pfi)
                if imgs:
                    all_packages.append({"name": pfi.name, "images": imgs})
            
            if not all_packages:
                st.error("No valid vendor pages could be prepared for OCR processing.")
            else:
                # Fire the multi-package extraction prompt
                extracted_data = extract_all_vendors_simultaneously(all_packages, row_item_dict, api_key)
                
                # Show live payload map for instant verification
                st.success("🤖 **Live Cross-Vendor Extraction Matrix Map:**")
                st.json(extracted_data)
                
                # Spreadsheet Target Tracks: E=Vendor 1 (Col 5), H=Vendor 2 (Col 8), K=Vendor 3 (Col 11), N=Vendor 4 (Col 14)
                vendor_columns = [5, 8, 11, 14]
                quotations_list = extracted_data.get("quotations", [])
                
                st.write("### ⚙️ Excel Multichannel Insertion Logs:")
                
                for q_data in quotations_list:
                    file_idx = q_data.get("file_index")
                    
                    # Ensure the AI returned an index matching our allocated tracks
                    if file_idx is None or file_idx >= len(vendor_columns):
                        continue
                        
                    target_col = vendor_columns[file_idx]
                    vendor_name = q_data.get("vendor_name", f"Vendor Slot {file_idx + 1}")
                    row_prices = q_data.get("row_prices", {})
                    
                    # 1. Place Vendor Name in Row 8 of their column tracking line
                    ws.cell(row=8, column=target_col).value = vendor_name
                    st.write(f"🏢 Assigned **{vendor_name}** header to Excel Column index `{target_col}`")
                    
                    # 2. Inject prices directly down the column matching the structural spreadsheet rows
                    match_count = 0
                    for row_str, price in row_prices.items():
                        try:
                            target_row = int(row_str)
                            clean_price = float(str(price).replace(",", "").strip())
                            
                            ws.cell(row=target_row, column=target_col).value = clean_price
                            match_count += 1
                        except (ValueError, TypeError):
                            pass
                            
                    st.success(f"🏁 Finished processing **{vendor_name}**. Injected {match_count} aligned items into Column Track `{target_col}`.")
                
                # Save and expose download
                output_filename = "Populated_Comparative_Analysis.xlsx"
                wb.save(output_filename)
                
                with open(output_filename, "rb") as file:
                    st.download_button(
                        label="📥 Download Ready Comparative Analysis Sheet",
                        data=file,
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                os.remove(output_filename)
